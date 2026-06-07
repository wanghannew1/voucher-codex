from pathlib import Path
import json
import re

import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

INPUT = Path('/Users/sunyitong/coding/voucher/代码资料/凭证信息_temp - 2026-06-02T140354.252.xlsx')
OUTDIR = Path('/Users/sunyitong/Documents/Codex/2026-06-06/new-chat/outputs/voucher_bank_analysis')
OUTDIR.mkdir(parents=True, exist_ok=True)

raw = pd.read_excel(INPUT, sheet_name='Sheet1', header=None, dtype=object)

# Row 1 has the machine-readable field list in A plus Chinese labels in B:BS.
field_codes = str(raw.iloc[1, 0]).strip().strip('"').split(',')
labels = [str(x).strip().replace('* ', '').replace('*', '') if pd.notna(x) else '' for x in raw.iloc[1, 1:1+len(field_codes)].tolist()]

data = raw.iloc[2:, :len(field_codes)].copy()
data.columns = field_codes
data = data.dropna(how='all').reset_index(drop=True)

for col in ['main_num', 'debitamount', 'localdebitamount', 'creditamount', 'localcreditamount']:
    if col in data.columns:
        data[col] = pd.to_numeric(data[col], errors='coerce')

for col in ['main_prepareddate', 'checkdate', 'verifydate']:
    if col in data.columns:
        data[col] = pd.to_datetime(data[col], errors='coerce')

data['accsubjcode_text'] = data['accsubjcode'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

def clean_key_part(value):
    if pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace('.0', '') if str(value).strip().endswith('.0') else str(value).strip()

data['voucher_key'] = (
    data['main_pk_accountingbook'].astype(str).fillna('') + '|' +
    data['main_pk_vouchertype'].astype(str).fillna('') + '|' +
    data['main_num'].apply(clean_key_part) + '|' +
    data['main_prepareddate'].dt.strftime('%Y-%m-%d').fillna('')
)

bank_line_mask = data['accsubjcode_text'].eq('1002')
bank_keys = set(data.loc[bank_line_mask, 'voucher_key'])
voucher_lines = data[data['voucher_key'].isin(bank_keys)].copy()
bank_lines = data[bank_line_mask].copy()

def signed_bank(row):
    debit = row.get('debitamount') if pd.notna(row.get('debitamount')) else 0
    credit = row.get('creditamount') if pd.notna(row.get('creditamount')) else 0
    return float(debit or 0) - float(credit or 0)

bank_lines['bank_net_amount'] = bank_lines.apply(signed_bank, axis=1)
bank_by_voucher = bank_lines.groupby('voucher_key', dropna=False).agg(
    日期=('main_prepareddate', 'first'),
    凭证号=('main_num', 'first'),
    银行借方金额=('debitamount', 'sum'),
    银行贷方金额=('creditamount', 'sum'),
    银行净流入=('bank_net_amount', 'sum'),
    银行分录数=('accsubjcode', 'count'),
    银行账户=('bankaccount', lambda s: '；'.join(sorted({str(x) for x in s.dropna() if str(x).strip() and str(x).lower() != 'nan'}))),
).reset_index()
bank_by_voucher['方向'] = bank_by_voucher['银行净流入'].apply(lambda x: '进账' if x > 0 else ('出账' if x < 0 else '平进平出/冲销'))
bank_by_voucher['金额绝对值'] = bank_by_voucher['银行净流入'].abs()

counterpart = voucher_lines[voucher_lines['accsubjcode_text'].ne('1002')].copy()
counterpart['amount_abs'] = counterpart[['debitamount', 'creditamount']].fillna(0).abs().sum(axis=1)
counter_summary = counterpart.groupby(['voucher_key', 'accsubjcode_text'], dropna=False).agg(
    对方科目借方=('debitamount', 'sum'),
    对方科目贷方=('creditamount', 'sum'),
    对方科目金额=('amount_abs', 'sum'),
    示例摘要=('explanation', lambda s: next((str(x) for x in s if pd.notna(x) and str(x).strip()), '')),
).reset_index().rename(columns={'accsubjcode_text': '对方科目编码'})

top_counter = counter_summary.merge(bank_by_voucher[['voucher_key', '方向']], on='voucher_key', how='left')
top_counter_summary = top_counter.groupby(['方向', '对方科目编码'], dropna=False).agg(
    凭证数=('voucher_key', 'nunique'),
    金额合计=('对方科目金额', 'sum'),
    借方合计=('对方科目借方', 'sum'),
    贷方合计=('对方科目贷方', 'sum'),
    示例摘要=('示例摘要', lambda s: next((str(x) for x in s if str(x).strip()), '')),
).reset_index().sort_values(['方向', '凭证数', '金额合计'], ascending=[True, False, False])

monthly = bank_by_voucher.assign(月份=bank_by_voucher['日期'].dt.to_period('M').astype(str)).groupby(['月份', '方向']).agg(
    凭证数=('voucher_key', 'nunique'),
    金额合计=('金额绝对值', 'sum'),
    平均金额=('金额绝对值', 'mean'),
).reset_index()

keyword_patterns = [
    ('工资/社保/公积金', r'工资|社保|公积金|薪'),
    ('税费', r'税|增值税|所得税|个税|附加'),
    ('收款/回款', r'收款|回款|到账|进账|客户|货款'),
    ('付款/报销', r'付款|支付|报销|转账|费用'),
    ('发票/调整/冲销', r'发票|调整|红冲|冲销|重开'),
    ('借款/往来', r'借款|往来|备用金|押金|保证金'),
]

voucher_text = voucher_lines.groupby('voucher_key')['explanation'].apply(lambda s: '；'.join(dict.fromkeys([str(x) for x in s if pd.notna(x) and str(x).strip()]))).reset_index(name='摘要合并')
bank_by_voucher = bank_by_voucher.merge(voucher_text, on='voucher_key', how='left')

def tag_text(text):
    text = str(text)
    tags = [name for name, pat in keyword_patterns if re.search(pat, text)]
    return '；'.join(tags) if tags else '其他'

bank_by_voucher['摘要类型'] = bank_by_voucher['摘要合并'].apply(tag_text)
keyword_summary = bank_by_voucher.groupby(['方向', '摘要类型']).agg(
    凭证数=('voucher_key', 'nunique'),
    金额合计=('金额绝对值', 'sum'),
    平均金额=('金额绝对值', 'mean'),
).reset_index().sort_values(['方向', '凭证数'], ascending=[True, False])

cols_keep = [
    'voucher_key', 'main_pk_accountingbook', 'main_pk_vouchertype', 'main_num', 'main_attachment',
    'main_pk_prepared', 'main_prepareddate', 'explanation', 'accsubjcode', 'pk_currtype',
    'debitamount', 'creditamount', 'bankaccount', 'billtype', 'ass_1', 'ass_2', 'unitname'
]
available = [c for c in cols_keep if c in voucher_lines.columns]

analysis_path = OUTDIR / '银行存款1002凭证筛选与规律分析.xlsx'
summary_rows = [
    ['源文件', str(INPUT)],
    ['原始分录数', len(data)],
    ['含1002银行存款的凭证数', bank_by_voucher['voucher_key'].nunique()],
    ['筛选后分录数（保留整张凭证）', len(voucher_lines)],
    ['1002银行存款分录数', len(bank_lines)],
    ['进账凭证数', int((bank_by_voucher['方向'] == '进账').sum())],
    ['出账凭证数', int((bank_by_voucher['方向'] == '出账').sum())],
    ['平进平出/冲销凭证数', int((bank_by_voucher['方向'] == '平进平出/冲销').sum())],
    ['进账金额合计', bank_by_voucher.loc[bank_by_voucher['方向'] == '进账', '金额绝对值'].sum()],
    ['出账金额合计', bank_by_voucher.loc[bank_by_voucher['方向'] == '出账', '金额绝对值'].sum()],
]

with pd.ExcelWriter(analysis_path, engine='openpyxl', datetime_format='yyyy-mm-dd') as writer:
    pd.DataFrame(summary_rows, columns=['指标', '数值']).to_excel(writer, sheet_name='分析摘要', index=False, startrow=1)
    bank_by_voucher.sort_values(['日期', '凭证号']).to_excel(writer, sheet_name='1002凭证清单', index=False)
    voucher_lines[available].sort_values(['main_prepareddate', 'main_num']).to_excel(writer, sheet_name='筛选后完整分录', index=False)
    top_counter_summary.to_excel(writer, sheet_name='对方科目规律', index=False)
    monthly.to_excel(writer, sheet_name='月度进出账', index=False)
    keyword_summary.to_excel(writer, sheet_name='摘要关键词规律', index=False)
    counter_summary.to_excel(writer, sheet_name='凭证对方科目明细', index=False)

    wb = writer.book
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='B7C9D6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        header_row = 2 if ws.title == '分析摘要' else 1
        ws.freeze_panes = f'A{header_row + 1}'
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 250), values_only=True):
                for value in cell:
                    max_len = max(max_len, len(str(value)) if value is not None else 0)
            width = min(max(max_len + 2, 12), 48)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' if abs(cell.value) != int(abs(cell.value)) else '#,##0'
                if hasattr(cell.value, 'year'):
                    cell.number_format = 'yyyy-mm-dd'
    ws = wb['分析摘要']
    ws['A1'] = '科目代码1002银行存款凭证筛选与规律分析'
    ws['A1'].font = Font(bold=True, size=14)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 72

result = {
    'output': str(analysis_path),
    'rows_total': int(len(data)),
    'voucher_count': int(bank_by_voucher['voucher_key'].nunique()),
    'line_count': int(len(voucher_lines)),
    'bank_line_count': int(len(bank_lines)),
    'direction_counts': bank_by_voucher['方向'].value_counts().to_dict(),
    'direction_amounts': bank_by_voucher.groupby('方向')['金额绝对值'].sum().round(2).to_dict(),
    'top_counter': top_counter_summary.head(15).to_dict(orient='records'),
    'keyword_summary': keyword_summary.to_dict(orient='records'),
    'monthly': monthly.to_dict(orient='records'),
}
print(json.dumps(result, ensure_ascii=False, default=str, indent=2))
